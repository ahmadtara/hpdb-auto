import streamlit as st
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
import re
from openpyxl import load_workbook
import pandas as pd

def run_boq():
    st.title("📊 KMZ ➜ BOQ - 1 FDT")

    kmz_file = st.file_uploader("Upload file .KMZ", type=["kmz"])
    template_file = st.file_uploader("Upload TEMPLATE BOQ - 1 FDT.xlsx", type=["xlsx"])

    if not (kmz_file and template_file):
        return

    kmz_bytes = kmz_file.read()

    # ---------- parse KML ----------
    def recurse_folder(folder, ns, path=""):
        items = []
        name_el = folder.find("kml:name", ns)
        folder_name = name_el.text.upper() if name_el is not None and name_el.text else "UNKNOWN"
        new_path = f"{path}/{folder_name}" if path else folder_name
        for sub in folder.findall("kml:Folder", ns):
            items += recurse_folder(sub, ns, new_path)
        for pm in folder.findall("kml:Placemark", ns):
            nm = pm.find("kml:name", ns)
            desc = pm.find("kml:description", ns)
            items.append({
                "name": nm.text.strip() if nm is not None and nm.text else "",
                "desc": desc.text.strip() if desc is not None and desc.text else "",
                "path": new_path
            })
        return items

    with zipfile.ZipFile(BytesIO(kmz_bytes)) as z:
        kml_files = [f for f in z.namelist() if f.lower().endswith(".kml")]
        if not kml_files:
            st.error("❌ Tidak ditemukan file .kml di dalam .kmz")
            return
        raw = z.read(kml_files[0])
        root = ET.fromstring(raw)
        ns = {"kml": "http://www.opengis.net/kml/2.2"}
        doc = root.find("kml:Document", ns) or root
        placemarks = recurse_folder(doc, ns, path="")

    # ---------- helper ----------
    def extract_number_from_string(s):
        if not s:
            return None
        m = re.search(r'[\d.,]+', s)
        if not m:
            return None
        token = m.group(0)
        parts = re.findall(r'\d+', token)
        if not parts:
            return None
        if len(parts) == 1:
            return int(parts[0])
        if len(parts) >= 2 and len(parts[-1]) == 3:
            return int(''.join(parts))
        return int(parts[0])

    def get_items(foldername):
        fn = foldername.upper()
        return [p for p in placemarks if fn in p["path"].split("/")]

    # ---------- collect ----------
    dist = get_items("DISTRIBUTION CABLE")
    sling = get_items("SLING WIRE")
    fat = get_items("FAT")
    hp_cover = get_items("HP COVER")

    lines = ["A", "B", "C", "D"]

    # ---------- Distribution ----------
    dist_per_line = {L: 0 for L in lines}
    for p in dist:
        for L in lines:
            if f"LINE {L}" in p["path"].split("/"):
                v = extract_number_from_string(p.get("desc", ""))
                if v is not None:
                    dist_per_line[L] += v
                break

    # ---------- Sling wire (A–D) ----------
    sling_items_abcd = [p for p in sling if any(f"LINE {L}" in p["path"].split("/") for L in lines)]
    sling_nums = []
    for p in sling_items_abcd:
        v = extract_number_from_string(p.get("name", ""))
        if v is not None:
            sling_nums.append(v)
    sling_total = sum(sling_nums)

    # ---------- FAT ----------
    fat_counts = {L: len([p for p in fat if f"LINE {L}" in p["path"].split("/")]) for L in lines}
    total_fat = sum(fat_counts.values())

    # ---------- HP COVER ----------
    hp_cover_counts = {L: len([p for p in hp_cover if f"LINE {L}" in p["path"].split("/")]) for L in lines}
    total_hp_cover = sum(hp_cover_counts.values())

    # ---------- Poles ----------
    def count_items_in_lines(items, lines_list=lines):
        return len([p for p in items if any(f"LINE {L}" in p["path"].split("/") for L in lines_list)])

    # Hitung total per tipe
    np74_items = get_items("NEW POLE 7-4")
    np73_items = get_items("NEW POLE 7-3")
    np725_items = get_items("NEW POLE 7-2.5")
    np94_items = get_items("NEW POLE 9-4")

    np74_count = count_items_in_lines(np74_items)
    np73_count = count_items_in_lines(np73_items)
    np725_count = count_items_in_lines(np725_items)
    np94_count = count_items_in_lines(np94_items)

    exist_pole = (
        get_items("EXISTING POLE EMR 7-4")
        + get_items("EXISTING POLE EMR 7-3")
        + get_items("EXISTING POLE EMR 7-2.5")
        + get_items("EXISTING POLE EMR 9-4")
    )
    exist_count = count_items_in_lines(exist_pole)

    # ---------- NEW: jumlah per LINE ----------
    def count_by_line(items):
        result = {}
        for L in lines:
            result[L] = len([p for p in items if f"LINE {L}" in p["path"].split("/")])
        return result

    np74_by_line = count_by_line(np74_items)
    np73_by_line = count_by_line(np73_items)
    np725_by_line = count_by_line(np725_items)
    np94_by_line = count_by_line(np94_items)
    exist_by_line = count_by_line(exist_pole)

    # ---------- Write to Excel ----------
    wb = load_workbook(template_file)
    ws = wb["BoM AE"]

    # Distribution mapping
    line_map = {"A": [2, 6, 10], "B": [3, 7, 11], "C": [4, 8, 12], "D": [5, 9, 13]}
    for line, cells in line_map.items():
        cnt = fat_counts.get(line, 0)
        val = dist_per_line.get(line, 0)
        if cnt >= 1 and cnt <= 10:
            ws[f"C{cells[0]}"] = val
        elif cnt >= 11 and cnt <= 15:
            ws[f"C{cells[1]}"] = val
        elif cnt >= 16 and cnt <= 20:
            ws[f"C{cells[2]}"] = val

    ws["C15"] = sling_total

    # FAT total -> C30..C32
    if 1 <= total_fat <= 24:
        ws["C30"] = 1
    elif 25 <= total_fat <= 36:
        ws["C31"] = 1
    elif 37 <= total_fat <= 48:
        ws["C32"] = 1

    ws["C36"] = fat_counts["A"]
    ws["C37"] = fat_counts["B"]
    ws["C38"] = fat_counts["C"]
    ws["C39"] = fat_counts["D"]

    # Poles (revisi final)
    ws["C54"] = np74_count
    ws["C55"] = np73_count
    ws["C56"] = np725_count
    ws["C58"] = np94_count
    ws["C61"] = exist_count

    # Sheet BoQ NRO Cluster
    ws2 = wb["BoQ NRO Cluster"]
    ws2["O5"] = total_hp_cover
    kmz_name = kmz_file.name.rsplit(".", 1)[0]
    ws2["O3"] = kmz_name

    # ---------- Output ----------
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # ---------- Streamlit output ----------
    st.subheader("📋 Summary from KMZ (raw counts)")
    df_summary = pd.DataFrame([
        ["Distribution A (desc sum)", dist_per_line["A"]],
        ["Distribution B (desc sum)", dist_per_line["B"]],
        ["Distribution C (desc sum)", dist_per_line["C"]],
        ["Distribution D (desc sum)", dist_per_line["D"]],
        ["Sling total", sling_total],
        ["FAT total (all lines)", total_fat],
        ["FAT A (count)", fat_counts["A"]],
        ["FAT B (count)", fat_counts["B"]],
        ["FAT C (count)", fat_counts["C"]],
        ["FAT D (count)", fat_counts["D"]],
        ["HP COVER total (all lines)", total_hp_cover],
        ["NEW POLE 7-4 total", np74_count],
        ["NEW POLE 7-3 total", np73_count],
        ["NEW POLE 7-2.5 total", np725_count],
        ["NEW POLE 9-4 total", np94_count],
        ["EXISTING POLE EMR total", exist_count],
    ], columns=["metric", "value"])
    st.dataframe(df_summary)

    st.subheader("📊 Pole per Line (A–D)")
    df_poles = pd.DataFrame({
        "Line": lines,
        "NEW POLE 7-4": [np74_by_line[L] for L in lines],
        "NEW POLE 7-3": [np73_by_line[L] for L in lines],
        "NEW POLE 7-2.5": [np725_by_line[L] for L in lines],
        "NEW POLE 9-4": [np94_by_line[L] for L in lines],
        "EXISTING POLE EMR": [exist_by_line[L] for L in lines],
    })
    st.dataframe(df_poles)

    st.success("✅ BOQ berhasil dibuat! (download di bawah)")
    st.download_button("📥 Download BOQ", buf.getvalue(), file_name="hasil_BOQ.xlsx")
